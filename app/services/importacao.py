from __future__ import annotations

import csv
import hashlib
import json
import re
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ..extensions import db
from ..models import ImportacaoLote, ImportacaoRegistro
from .processos import normalizar_codigo


ARQUIVOS = {
    "agenda.CSV": {"Processo", "Nome do cliente", "Tipo", "Status"},
    "servicos.CSV": {"Processo", "Cliente", "Serviços", "Status"},
    "avaliacao_bruta.CSV": {"Processo Nº", "Cliente", "Data"},
}

DATAS = {
    "agenda.CSV": (
        "Recebido em", "Data A inicial", "Data B inicial",
        "Data A ofertada", "Data B ofertada",
    ),
    "servicos.CSV": ("data_inicio", "data_final"),
    "avaliacao_bruta.CSV": ("Data",),
}

NOTAS = (
    "Pontualidade/ Coordenação",
    "Limpeza/ Qualidade de Embalagem",
    "Cortesia/ Qualidade de Carregamento",
    "Técnica/ Cortesia",
)

DOMINIOS = {
    ("agenda.CSV", "Status"): {"Confirmado", "Não", "Pré-agendado"},
    ("servicos.CSV", "Serviços"): {
        "Exportação", "Importação", "Doméstico", "Guarda Móveis",
    },
    ("servicos.CSV", "Status"): {"Finalizado", "Agendado"},
    ("servicos.CSV", "Modal"): {"SEA", "AIR", "LAND", "GM"},
}


class ErroImportacao(ValueError):
    pass


@dataclass(frozen=True)
class Linha:
    arquivo: str
    numero: int
    bruto: dict
    valores: dict[str, str]
    checksum: str


def normalizar_cabecalho(valor: str) -> str:
    return " ".join(str(valor).split())


def _ler_arquivo(caminho: Path) -> list[Linha]:
    try:
        arquivo = caminho.open(encoding="latin-1", newline="")
    except OSError as exc:
        raise ErroImportacao(f"Não foi possível ler {caminho}: {exc}") from exc
    with arquivo:
        leitor = csv.DictReader(arquivo, delimiter=";")
        if not leitor.fieldnames:
            raise ErroImportacao(f"{caminho.name} não possui cabeçalho")
        presentes = {normalizar_cabecalho(c) for c in leitor.fieldnames if c}
        ausentes = ARQUIVOS[caminho.name] - presentes
        if ausentes:
            raise ErroImportacao(
                f"{caminho.name}: cabeçalhos obrigatórios ausentes: "
                + ", ".join(sorted(ausentes))
            )
        linhas = []
        for numero, row in enumerate(leitor, start=2):
            bruto = {
                (str(chave) if chave is not None else "__extra__"): valor
                for chave, valor in row.items()
            }
            valores = {
                normalizar_cabecalho(chave): str(valor or "").strip()
                for chave, valor in row.items() if chave is not None
            }
            checksum = hashlib.sha256(
                json.dumps(bruto, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
            ).hexdigest()
            linhas.append(Linha(caminho.name, numero, bruto, valores, checksum))
        return linhas


def ler_fontes(diretorio: Path) -> tuple[str, list[Linha]]:
    diretorio = Path(diretorio)
    if not diretorio.is_dir():
        raise ErroImportacao(f"Diretório não encontrado: {diretorio}")
    nomes = {item.name.lower(): item for item in diretorio.iterdir() if item.is_file()}
    caminhos = []
    for nome in ARQUIVOS:
        caminho = nomes.get(nome.lower())
        if not caminho:
            raise ErroImportacao(f"Arquivo obrigatório ausente: {nome}")
        caminhos.append((nome, caminho))

    digest = hashlib.sha256()
    linhas = []
    for nome, caminho in caminhos:
        digest.update(nome.encode())
        digest.update(b"\0")
        digest.update(hashlib.sha256(caminho.read_bytes()).digest())
        linhas.extend(_ler_arquivo(caminho))
    return digest.hexdigest(), linhas


def normalizar_intervalo(
    valor: str, ano_referencia: int | None = None, formato_americano: bool = False
) -> tuple[date, date] | None:
    valor = valor.strip()
    if not valor:
        return None

    def converter(texto: str, ano_padrao=None, americano=False):
        texto = texto.strip()
        if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", texto):
            return date.fromisoformat(texto)
        partes = texto.split("/")
        if len(partes) not in (2, 3):
            raise ValueError
        primeiro, segundo = (int(partes[0]), int(partes[1]))
        ano = int(partes[2]) if len(partes) == 3 else ano_padrao
        if ano is None:
            raise ValueError
        usar_americano = americano and primeiro <= 12
        dia, mes = (segundo, primeiro) if usar_americano else (primeiro, segundo)
        return date(ano, mes, dia)

    try:
        faixa = re.fullmatch(
            r"(\d{1,2}/\d{1,2}(?:/\d{4})?)\s+a\s+"
            r"(\d{1,2}/\d{1,2}(?:/\d{4})?)", valor, re.I,
        )
        if faixa:
            inicio = converter(faixa.group(1), ano_referencia)
            fim = converter(faixa.group(2), inicio.year)
            if fim < inicio:
                fim = fim.replace(year=fim.year + 1)
            return inicio, fim
        dias = re.fullmatch(r"(\d{1,2})\s+(?:e|a)\s+(\d{1,2})/(\d{1,2})(?:/(\d{4}))?", valor, re.I)
        if dias:
            ano = int(dias.group(4)) if dias.group(4) else ano_referencia
            if ano is None:
                raise ValueError
            inicio = date(ano, int(dias.group(3)), int(dias.group(1)))
            fim = date(ano, int(dias.group(3)), int(dias.group(2)))
            return inicio, fim
        unica = converter(valor, ano_referencia, americano=formato_americano)
        return unica, unica
    except (ValueError, TypeError):
        return None


def _decimal(valor: str) -> Decimal | None:
    if not valor:
        return None
    try:
        return Decimal(valor.replace(".", "").replace(",", "."))
    except InvalidOperation:
        return None


def analisar(linhas: list[Linha]) -> dict:
    problemas = []
    por_linha = defaultdict(list)

    def adicionar(linha, codigo, campo, valor, mensagem, severidade="aviso"):
        item = {
            "severidade": severidade,
            "codigo": codigo,
            "arquivo": linha.arquivo if linha else "",
            "linha": linha.numero if linha else "",
            "processo": _processo(linha) if linha else "",
            "campo": campo,
            "valor_bruto": valor,
            "mensagem": mensagem,
        }
        problemas.append(item)
        if linha:
            por_linha[(linha.arquivo, linha.numero)].append(codigo)

    def linhas_de(arquivo):
        return [linha for linha in linhas if linha.arquivo == arquivo]

    for linha in linhas:
        processo = _processo(linha)
        financeiro = linha.arquivo == "avaliacao_bruta.CSV" and not processo
        if not processo and linha.arquivo == "servicos.CSV":
            adicionar(
                linha, "processo_ausente", _campo_processo(linha.arquivo), "",
                "A linha não possui identificador de processo.", "erro",
            )
        recebido = normalizar_intervalo(
            linha.valores.get("Recebido em", ""), formato_americano=True
        ) if linha.arquivo == "agenda.CSV" else None
        ano_referencia = recebido[0].year if recebido else None
        for campo in DATAS[linha.arquivo]:
            valor = linha.valores.get(campo, "")
            americano = linha.arquivo == "agenda.CSV" and campo == "Recebido em"
            if valor and not normalizar_intervalo(valor, ano_referencia, americano):
                adicionar(
                    linha, "data_invalida", campo, valor,
                    "O valor não pôde ser convertido para intervalo em dd/mm/aaaa.", "erro",
                )
        for (arquivo, campo), permitidos in DOMINIOS.items():
            if linha.arquivo != arquivo:
                continue
            valor = linha.valores.get(campo, "")
            if valor and valor not in permitidos:
                adicionar(
                    linha, "dominio_desconhecido", campo, valor,
                    "Valor ainda não possui regra de mapeamento.",
                )

        if linha.arquivo == "avaliacao_bruta.CSV" and not financeiro:
            for campo in NOTAS:
                valor = linha.valores.get(campo, "")
                nota = _decimal(valor)
                if valor and (nota is None or not Decimal("0") <= nota <= Decimal("10")):
                    adicionar(
                        linha, "nota_invalida", campo, valor,
                        "A nota deve ser numérica e estar entre 0 e 10.", "erro",
                    )

    codigos = defaultdict(lambda: defaultdict(list))
    for linha in linhas:
        bruto = _processo(linha)
        if bruto:
            codigos[normalizar_codigo(bruto)][bruto].append(linha)
    for normalizado, variantes in codigos.items():
        if len(variantes) > 1:
            for bruto, ocorrencias in variantes.items():
                adicionar(
                    ocorrencias[0], "colisao_codigo", _campo_processo(ocorrencias[0].arquivo),
                    bruto, f"Colide com outro código após normalização para {normalizado}.", "erro",
                )

    agenda = {
        normalizar_codigo(_processo(linha))
        for linha in linhas_de("agenda.CSV") if _processo(linha)
    }
    outras = defaultdict(list)
    for linha in linhas:
        bruto = _processo(linha)
        if linha.arquivo != "agenda.CSV" and bruto:
            outras[normalizar_codigo(bruto)].append(linha)

    orfaos = []
    for codigo in sorted(set(outras) - agenda):
        ocorrencias = outras[codigo]
        fontes = sorted({linha.arquivo for linha in ocorrencias})
        clientes = sorted({linha.valores.get("Cliente", "") for linha in ocorrencias if linha.valores.get("Cliente", "")})
        tipos = sorted({linha.valores.get("Serviços", "") for linha in ocorrencias if linha.valores.get("Serviços", "")})
        primeira = ocorrencias[0]
        orfaos.append({
            "processo": codigo,
            "fontes": ", ".join(fontes),
            "cliente_candidato": " | ".join(clientes),
            "tipo_candidato": " | ".join(tipos),
            "ocorrencias": len(ocorrencias),
        })
        adicionar(
            primeira, "processo_orfao", _campo_processo(primeira.arquivo),
            _processo(primeira),
            "O processo não aparece na agenda; as tabelas antigas não eram usadas continuamente.",
            "aviso",
        )

    solicitacoes = []
    movimentos_financeiros = []
    datas_normalizadas = []
    for linha in linhas:
        processo = _processo(linha)
        if linha.arquivo == "agenda.CSV" and not processo:
            recebido = normalizar_intervalo(linha.valores.get("Recebido em", ""), formato_americano=True)
            ano = recebido[0].year if recebido else None
            solicitacoes.append({
                "arquivo": linha.arquivo,
                "linha": linha.numero,
                "cliente": linha.valores.get("Nome do cliente", ""),
                "origem": linha.valores.get("Origem", ""),
                "destino": linha.valores.get("Destino", ""),
                "volume": linha.valores.get("Volume", ""),
                "data_inicial": _intervalo_texto(normalizar_intervalo(linha.valores.get("Data A inicial", ""), ano)),
                "data_ofertada": _intervalo_texto(normalizar_intervalo(linha.valores.get("Data A ofertada", ""), ano)),
                "status": linha.valores.get("Status", ""),
            })
        elif linha.arquivo == "avaliacao_bruta.CSV" and not processo:
            movimentos_financeiros.append({
                "arquivo": linha.arquivo,
                "linha": linha.numero,
                "data": linha.valores.get("Data", ""),
                "descricao": linha.valores.get("Cliente", ""),
                "valor": linha.valores.get("$ total", ""),
                "comentario": linha.valores.get("Comentário", ""),
            })
        ano = None
        if linha.arquivo == "agenda.CSV":
            recebido = normalizar_intervalo(linha.valores.get("Recebido em", ""), formato_americano=True)
            ano = recebido[0].year if recebido else None
        for campo in DATAS[linha.arquivo]:
            valor = linha.valores.get(campo, "")
            intervalo = normalizar_intervalo(
                valor, ano, linha.arquivo == "agenda.CSV" and campo == "Recebido em"
            ) if valor else None
            if intervalo:
                datas_normalizadas.append({
                    "arquivo": linha.arquivo, "linha": linha.numero,
                    "processo": processo, "campo": campo, "valor_bruto": valor,
                    "inicio": intervalo[0].strftime("%d/%m/%Y"),
                    "fim": intervalo[1].strftime("%d/%m/%Y"),
                })

    dominios = []
    campos = {
        "agenda.CSV": ("Tipo", "Status"),
        "servicos.CSV": ("Serviços", "Modal", "Status", "Private/Booking", "Coordenadora", "Equipe"),
    }
    for arquivo, nomes in campos.items():
        registros = linhas_de(arquivo)
        for campo in nomes:
            contagem = Counter(linha.valores.get(campo, "") for linha in registros)
            for valor, quantidade in sorted(contagem.items(), key=lambda x: (-x[1], x[0])):
                if valor:
                    dominios.append({
                        "arquivo": arquivo,
                        "campo": campo,
                        "valor": valor,
                        "valor_normalizado": primeira_coordenadora(valor) if campo == "Coordenadora" else "",
                        "quantidade": quantidade,
                    })

    totais_arquivo = Counter(linha.arquivo for linha in linhas)
    severidades = Counter(item["severidade"] for item in problemas)
    return {
        "problemas": problemas,
        "por_linha": dict(por_linha),
        "orfaos": orfaos,
        "solicitacoes": solicitacoes,
        "movimentos_financeiros": movimentos_financeiros,
        "datas_normalizadas": datas_normalizadas,
        "dominios": dominios,
        "totais": {
            "linhas": len(linhas),
            "arquivos": dict(totais_arquivo),
            "problemas": len(problemas),
            "erros": severidades["erro"],
            "avisos": severidades["aviso"],
            "processos_orfaos": len(orfaos),
            "solicitacoes": len(solicitacoes),
            "movimentos_financeiros": len(movimentos_financeiros),
        },
    }


def _intervalo_texto(intervalo):
    if not intervalo:
        return ""
    return f"{intervalo[0].strftime('%d/%m/%Y')} a {intervalo[1].strftime('%d/%m/%Y')}"


def primeira_coordenadora(valor: str) -> str:
    return re.split(r"\s*(?:,|;|\s+e\s+)\s*", valor, maxsplit=1, flags=re.I)[0].strip()


def _campo_processo(arquivo: str) -> str:
    return "Processo Nº" if arquivo == "avaliacao_bruta.CSV" else "Processo"


def _processo(linha: Linha) -> str:
    return linha.valores.get(_campo_processo(linha.arquivo), "").strip()


def carregar(diretorio: Path, dry_run: bool = False) -> dict:
    fingerprint, linhas = ler_fontes(Path(diretorio))
    diagnostico = analisar(linhas)
    lote_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"filial-bsb:{fingerprint}"))
    resultado = {
        "lote": lote_id,
        "fingerprint": fingerprint,
        "dry_run": dry_run,
        "reutilizado": False,
        **diagnostico["totais"],
    }
    if dry_run:
        return resultado

    existente = ImportacaoLote.query.filter_by(fingerprint=fingerprint).first()
    if existente:
        resultado["lote"] = existente.id
        resultado["reutilizado"] = True
        resultado.update(existente.totais)
        return resultado

    lote = ImportacaoLote(
        id=lote_id,
        fingerprint=fingerprint,
        origem=str(Path(diretorio).resolve()),
        estado="diagnosticado",
        totais=diagnostico["totais"],
    )
    db.session.add(lote)
    for linha in linhas:
        codigos = diagnostico["por_linha"].get((linha.arquivo, linha.numero), [])
        db.session.add(ImportacaoRegistro(
            lote=lote.id,
            arquivo=linha.arquivo,
            aba=None,
            linha=linha.numero,
            checksum=linha.checksum,
            valor_bruto=linha.bruto,
            estado="conflito" if codigos else "carregado",
            conflito=", ".join(sorted(set(codigos))) or None,
        ))
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    return resultado


def linhas_do_lote(lote_id: str) -> tuple[ImportacaoLote, list[Linha]]:
    lote = db.session.get(ImportacaoLote, lote_id)
    if not lote:
        raise ErroImportacao(f"Lote não encontrado: {lote_id}")
    registros = ImportacaoRegistro.query.filter_by(lote=lote.id).order_by(
        ImportacaoRegistro.arquivo, ImportacaoRegistro.linha
    ).all()
    linhas = []
    for registro in registros:
        valores = {
            normalizar_cabecalho(chave): str(valor or "").strip()
            for chave, valor in registro.valor_bruto.items() if chave != "__extra__"
        }
        linhas.append(Linha(
            registro.arquivo, registro.linha, registro.valor_bruto,
            valores, registro.checksum,
        ))
    return lote, linhas


def gerar_relatorio(lote_id: str, destino: Path) -> dict:
    lote, linhas = linhas_do_lote(lote_id)
    diagnostico = analisar(linhas)
    wb = Workbook()
    wb.remove(wb.active)

    resumo = wb.create_sheet("Resumo")
    resumo.append(["Indicador", "Valor"])
    resumo.append(["Lote", lote.id])
    resumo.append(["Fingerprint", lote.fingerprint])
    resumo.append(["Origem", lote.origem])
    resumo.append(["Gerado em", datetime.now().isoformat(timespec="seconds")])
    for nome, valor in diagnostico["totais"].items():
        if nome != "arquivos":
            resumo.append([nome, valor])
    for arquivo, quantidade in diagnostico["totais"]["arquivos"].items():
        resumo.append([f"linhas:{arquivo}", quantidade])

    _adicionar_tabela(wb, "Pendencias", diagnostico["problemas"], (
        "severidade", "codigo", "arquivo", "linha", "processo",
        "campo", "valor_bruto", "mensagem",
    ))
    _adicionar_tabela(wb, "Processos_sem_agenda", diagnostico["orfaos"], (
        "processo", "fontes", "cliente_candidato", "tipo_candidato", "ocorrencias",
    ))
    _adicionar_tabela(wb, "Dominios", diagnostico["dominios"], (
        "arquivo", "campo", "valor", "valor_normalizado", "quantidade",
    ))
    _adicionar_tabela(wb, "Solicitacoes", diagnostico["solicitacoes"], (
        "arquivo", "linha", "cliente", "origem", "destino", "volume",
        "data_inicial", "data_ofertada", "status",
    ))
    _adicionar_tabela(wb, "Registros_financeiros", diagnostico["movimentos_financeiros"], (
        "arquivo", "linha", "data", "descricao", "valor", "comentario",
    ))
    _adicionar_tabela(wb, "Datas_normalizadas", diagnostico["datas_normalizadas"], (
        "arquivo", "linha", "processo", "campo", "valor_bruto", "inicio", "fim",
    ))

    for arquivo in ARQUIVOS:
        registros = [linha for linha in linhas if linha.arquivo == arquivo]
        cabecalhos = []
        for linha in registros:
            for chave in linha.bruto:
                if chave not in cabecalhos:
                    cabecalhos.append(chave)
        ws = wb.create_sheet({
            "agenda.CSV": "Bruto_agenda",
            "servicos.CSV": "Bruto_servicos",
            "avaliacao_bruta.CSV": "Bruto_avaliacoes",
        }[arquivo])
        ws.append(["linha", *cabecalhos])
        for linha in registros:
            ws.append([linha.numero, *[_excel(linha.bruto.get(campo, "")) for campo in cabecalhos]])
        _formatar(ws)

    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return diagnostico["totais"]


def _excel(valor):
    if isinstance(valor, (str, int, float, date)) or valor is None:
        return valor
    return json.dumps(valor, ensure_ascii=False)


def _adicionar_tabela(wb, titulo, itens, campos):
    ws = wb.create_sheet(titulo)
    ws.append(list(campos))
    for item in itens:
        ws.append([_excel(item.get(campo, "")) for campo in campos])
    _formatar(ws)


def _formatar(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for coluna in range(1, min(ws.max_column, 20) + 1):
        largura = max(
            (len(str(ws.cell(linha, coluna).value or "")) for linha in range(1, min(ws.max_row, 200) + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(coluna)].width = min(max(largura + 2, 10), 48)
