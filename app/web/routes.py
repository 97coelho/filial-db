import csv
import io
from datetime import date
from functools import wraps

from flask import Blueprint, Response, flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font

from ..extensions import db
from ..models import Etapa, MovimentoPool, Processo, Usuario
from ..services.processos import buscar_processos, pendencias

web = Blueprint("web", __name__)


def login_required(fn):
    @wraps(fn)
    def wrapped(*args, **kwargs):
        return fn(*args, **kwargs) if session.get("usuario_id") else redirect(url_for("web.login", next=request.path))
    return wrapped


@web.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = Usuario.query.filter_by(email=request.form.get("email", "").lower(), ativo=True).first()
        if user and user.check_password(request.form.get("senha", "")):
            session.clear(); session["usuario_id"] = user.id
            return redirect(request.args.get("next") or url_for("web.dashboard"))
        flash("E-mail ou senha inválidos.", "erro")
    return render_template("login.html")


@web.post("/logout")
def logout():
    session.clear(); return redirect(url_for("web.login"))


@web.get("/")
@login_required
def dashboard():
    processos = Processo.query.order_by(Processo.criado_em.desc()).limit(10).all()
    proximos = Processo.query.join(Processo.janelas).filter_by(tipo="confirmada").filter(db.text("janelas_agenda.data_fim >= CURRENT_DATE")).limit(10).all()
    vencidas = Etapa.query.filter(Etapa.prevista_fim < date.today(), Etapa.status != "concluida").count()
    aguardando = MovimentoPool.query.filter_by(estado="calculado").count()
    return render_template("dashboard.html", processos=processos, proximos=proximos, vencidas=vencidas,
                           aguardando=aguardando, criticos=sum(bool(pendencias(p)) for p in Processo.query.all()))


@web.get("/processos")
@login_required
def processos():
    items = buscar_processos(request.args.get("q"), request.args.get("status_agenda"), request.args.get("status_operacao")).all()
    return render_template("processos.html", processos=items, filtros=request.args)


@web.get("/processos/<uuid>")
@login_required
def processo(uuid):
    item = db.get_or_404(Processo, uuid)
    return render_template("processo.html", processo=item, pendencias=pendencias(item))


def rows():
    return [[p.codigo, p.cliente.nome, p.tipo, p.status_agenda, p.status_operacao, p.criado_em.date()] for p in
            buscar_processos(request.args.get("q"), request.args.get("status_agenda"), request.args.get("status_operacao")).all()]


@web.get("/exportar/processos.csv")
@login_required
def export_csv():
    output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow(["Código", "Cliente", "Tipo", "Agenda", "Operação", "Criado em"]); writer.writerows(rows())
    return Response(output.getvalue(), mimetype="text/csv; charset=utf-8", headers={"Content-Disposition": "attachment; filename=processos.csv"})


@web.get("/exportar/processos.xlsx")
@login_required
def export_xlsx():
    wb = Workbook(); ws = wb.active; ws.title = "Processos"
    headers = ["Código", "Cliente", "Tipo", "Agenda", "Operação", "Criado em"]; ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True)
    for row in rows(): ws.append(row)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for cell in ws["F"][1:]: cell.number_format = "dd/mm/yyyy"
    for col, width in zip("ABCDEF", [16, 32, 18, 18, 18, 14]): ws.column_dimensions[col].width = width
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="processos.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
