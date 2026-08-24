import csv
from datetime import date

from openpyxl import load_workbook

from app.extensions import db
from app.models import AvaliacaoTemplate, ImportacaoLote, ImportacaoRegistro, Pessoa, Processo
from app.seed import seed_database
from app.services.importacao import (
    analisar, carregar, gerar_relatorio, ler_fontes, normalizar_intervalo,
    primeira_coordenadora,
)


def escrever_csv(caminho, cabecalhos, linhas):
    with caminho.open("w", encoding="latin-1", newline="") as arquivo:
        writer = csv.DictWriter(arquivo, fieldnames=cabecalhos, delimiter=";")
        writer.writeheader()
        writer.writerows(linhas)


def fontes(tmp_path, cliente="Maria", nota="11"):
    escrever_csv(
        tmp_path / "agenda.CSV",
        ["Processo", "Nome do cliente", "Tipo", "Status", "Recebido em", "Data A inicial"],
        [{
            "Processo": " EXP-1 ", "Nome do cliente": cliente, "Tipo": "Embalagem",
            "Status": "Confirmado", "Recebido em": "20/08/2026", "Data A inicial": "ASAP",
        }],
    )
    escrever_csv(
        tmp_path / "servicos.CSV",
        ["Processo", "Cliente", "Serviços", "Status", "Modal", "Coordenadora", "Equipe", "data_inicio"],
        [
            {"Processo": "EXP-1", "Cliente": cliente, "Serviços": "Exportação", "Status": "Finalizado", "Modal": "SEA", "Coordenadora": "Ana", "Equipe": "João", "data_inicio": "20/08/2026"},
            {"Processo": "IMP-2", "Cliente": "José", "Serviços": "Importação", "Status": "Finalizado", "Modal": "AIR", "Coordenadora": "Ana, Bia", "Equipe": "João e Bia", "data_inicio": "21/08/2026"},
        ],
    )
    escrever_csv(
        tmp_path / "avaliacao_bruta.CSV",
        ["Processo Nº", "Cliente", "Data", "Pontualidade/\nCoordenação"],
        [{"Processo Nº": "IMP-2", "Cliente": "José", "Data": "22/08/2026", "Pontualidade/\nCoordenação": nota}],
    )


def test_dry_run_nao_grava_e_identifica_problemas(app, tmp_path):
    fontes(tmp_path)
    with app.app_context():
        resultado = carregar(tmp_path, dry_run=True)
        assert resultado["linhas"] == 4
        assert resultado["processos_orfaos"] == 1
        assert resultado["erros"] >= 2
        assert ImportacaoLote.query.count() == 0
        assert ImportacaoRegistro.query.count() == 0


def test_carga_idempotente_e_novo_snapshot(app, tmp_path):
    fontes(tmp_path)
    with app.app_context():
        primeiro = carregar(tmp_path)
        segundo = carregar(tmp_path)
        assert segundo["lote"] == primeiro["lote"]
        assert segundo["reutilizado"] is True
        assert ImportacaoLote.query.count() == 1
        assert ImportacaoRegistro.query.count() == 4

        fontes(tmp_path, cliente="Maria Silva")
        terceiro = carregar(tmp_path)
        assert terceiro["lote"] != primeiro["lote"]
        assert ImportacaoLote.query.count() == 2
        assert ImportacaoRegistro.query.count() == 8
        assert Pessoa.query.count() == 0
        assert Processo.query.count() == 0


def test_relatorio_xlsx_tem_diagnostico_e_dados_brutos(app, tmp_path):
    fontes(tmp_path)
    destino = tmp_path / "reports" / "diagnostico.xlsx"
    with app.app_context():
        resultado = carregar(tmp_path)
        totais = gerar_relatorio(resultado["lote"], destino)
    wb = load_workbook(destino)
    assert set((
        "Resumo", "Pendencias", "Processos_sem_agenda", "Dominios",
        "Solicitacoes", "Registros_financeiros", "Datas_normalizadas",
        "Bruto_agenda", "Bruto_servicos", "Bruto_avaliacoes",
    )).issubset(wb.sheetnames)
    assert totais["linhas"] == 4
    assert wb["Processos_sem_agenda"].max_row == 2


def test_cli_rejeita_fonte_incompleta(app, tmp_path):
    runner = app.test_cli_runner()
    resultado = runner.invoke(args=["importar", "carregar", str(tmp_path), "--dry-run"])
    assert resultado.exit_code != 0
    assert "Arquivo obrigatório ausente" in resultado.output


def test_notas_de_zero_a_dez_sao_validas(tmp_path):
    for nota in ("0", "10"):
        fontes(tmp_path, nota=nota)
        _, linhas = ler_fontes(tmp_path)
        diagnostico = analisar(linhas)
        assert not any(
            item["codigo"] == "nota_invalida" for item in diagnostico["problemas"]
        )


def test_seed_atualiza_escala_do_template_existente(app):
    with app.app_context():
        template = AvaliacaoTemplate(
            nome="Avaliação padrão", versao=1, vigente_desde=date(2024, 1, 1),
            perguntas=[{"codigo": "nota_1", "escala": [1, 5]}],
        )
        db.session.add(template)
        db.session.commit()
        seed_database()
        assert all(pergunta["escala"] == [0, 10] for pergunta in template.perguntas)


def test_normaliza_datas_e_intervalos_com_ano_de_referencia():
    assert normalizar_intervalo("3/20/2026", formato_americano=True) == (
        date(2026, 3, 20), date(2026, 3, 20),
    )
    assert normalizar_intervalo("30/03 a 02/04", 2026) == (
        date(2026, 3, 30), date(2026, 4, 2),
    )
    assert normalizar_intervalo("19 e 20/01", 2026) == (
        date(2026, 1, 19), date(2026, 1, 20),
    )
    assert normalizar_intervalo("02 a 06/03", 2026) == (
        date(2026, 3, 2), date(2026, 3, 6),
    )
    assert normalizar_intervalo("16/03/2026", formato_americano=True) == (
        date(2026, 3, 16), date(2026, 3, 16),
    )
    assert normalizar_intervalo("ASAP", 2026) is None


def test_coordenadora_composta_usa_primeiro_nome():
    assert primeira_coordenadora("Leticia, Gabrielle") == "Leticia"
    assert primeira_coordenadora("Ana e Bianca") == "Ana"
