"""
gerar_template.py
------------------
Gera a planilha `template_migracao_dados.xlsx`, com uma aba por tabela do
banco `mudancas_filial.db`, para preenchimento manual dos ~110 processos.

Estrutura de cada aba de dados (Agenda, Clientes, Servicos, Servicos_Equipe,
Avaliacoes_Brutas):
  Linha 1 -> Cabeçalho (nomes de coluna, iguais ao schema do banco)
  Linha 2 -> DICA de formato/tipo esperado (ex: "Número inteiro (ex.: 15000)")
  Linha 3 -> EXEMPLO com valores realistas preenchidos
  Linha 4+ -> onde você preenche os dados de verdade

Aba especial "Dicionario": guarda as listas de valores padronizados
(cidade, empresa, agente, coordenadora, status, modal, etc.). As listas
suspensas das outras abas apontam para as colunas dessa aba — para
adicionar uma opção nova, basta digitar na próxima linha vazia da coluna
correspondente no Dicionario, sem mexer em nenhuma fórmula.

Rodar com: python3 gerar_template.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONTE = "Arial"
AZUL_ESCURO = "1F4E78"
CINZA_DICA = "F2F2F2"
AMARELO_EXEMPLO = "FFF2CC"

LINHA_CABECALHO = 1
LINHA_DICA = 2
LINHA_EXEMPLO = 3
PRIMEIRA_LINHA_DADOS = 4
ULTIMA_LINHA_VALIDACAO = 300  # margem confortável acima dos ~110 processos

wb = Workbook()
wb.remove(wb.active)  # remove a aba padrão em branco


# =====================================================================
# Funções auxiliares de estilo
# =====================================================================
def estilizar_cabecalho(ws, num_colunas):
    for col in range(1, num_colunas + 1):
        celula = ws.cell(row=LINHA_CABECALHO, column=col)
        celula.font = Font(name=FONTE, bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[LINHA_CABECALHO].height = 30


def estilizar_linha_dica(ws, num_colunas):
    for col in range(1, num_colunas + 1):
        celula = ws.cell(row=LINHA_DICA, column=col)
        celula.font = Font(name=FONTE, italic=True, size=9, color="595959")
        celula.fill = PatternFill("solid", fgColor=CINZA_DICA)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[LINHA_DICA].height = 28


def estilizar_linha_exemplo(ws, num_colunas, linha=LINHA_EXEMPLO):
    for col in range(1, num_colunas + 1):
        celula = ws.cell(row=linha, column=col)
        celula.font = Font(name=FONTE, italic=True, color="7F6000")
        celula.fill = PatternFill("solid", fgColor=AMARELO_EXEMPLO)


def montar_aba_dados(nome_aba, especificacoes):
    """
    especificacoes: lista de tuplas (nome_coluna, dica, valor_exemplo)
    Monta cabeçalho + linha de dica + linha de exemplo de uma só vez,
    para não repetir esse bloco em cada aba.
    """
    ws = wb.create_sheet(nome_aba)
    colunas = [c[0] for c in especificacoes]
    dicas = [c[1] for c in especificacoes]
    exemplos = [c[2] for c in especificacoes]

    ws.append(colunas)
    ws.append(dicas)
    ws.append(exemplos)

    estilizar_cabecalho(ws, len(colunas))
    estilizar_linha_dica(ws, len(colunas))
    estilizar_linha_exemplo(ws, len(colunas))
    ws.freeze_panes = f"A{PRIMEIRA_LINHA_DADOS}"  # cabeçalho+dica+exemplo sempre visíveis
    for i, _ in enumerate(colunas, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    return ws, colunas


def aplicar_formato_data(ws, colunas, nomes_colunas_data):
    for nome in nomes_colunas_data:
        idx = colunas.index(nome) + 1
        letra = get_column_letter(idx)
        for linha in range(PRIMEIRA_LINHA_DADOS, ULTIMA_LINHA_VALIDACAO + 1):
            ws[f"{letra}{linha}"].number_format = "yyyy-mm-dd"


def adicionar_dropdown_dicionario(ws_alvo, colunas, nome_coluna, letra_dicionario):
    """
    Cria uma lista suspensa que aponta para uma coluna da aba Dicionario
    (a partir da linha 3 dela, onde os valores começam), em vez de uma
    lista fixa escrita na própria fórmula. Assim, para adicionar uma
    opção nova (ex: uma cidade nova), basta editar o Dicionario — o
    dropdown já enxerga o valor novo automaticamente.
    """
    idx = colunas.index(nome_coluna) + 1
    letra = get_column_letter(idx)
    dv = DataValidation(
        type="list",
        formula1=f"=Dicionario!${letra_dicionario}$3:${letra_dicionario}${ULTIMA_LINHA_VALIDACAO}",
        allow_blank=True,
        showDropDown=False,  # False = mostra a seta do dropdown (comportamento do Excel é invertido aqui)
    )
    ws_alvo.add_data_validation(dv)
    dv.add(f"{letra}{PRIMEIRA_LINHA_DADOS}:{letra}{ULTIMA_LINHA_VALIDACAO}")


# =====================================================================
# Aba: Instrucoes
# =====================================================================
ws = wb.create_sheet("Instrucoes")
ws.column_dimensions["A"].width = 100
instrucoes = [
    ("COMO USAR ESTA PLANILHA", True),
    ("", False),
    ("1. Abas de dados: Agenda, Clientes, Servicos, Servicos_Equipe e "
     "Avaliacoes_Brutas — cada uma corresponde a uma tabela do banco "
     "mudancas_filial.db.", False),
    ("2. Cada aba de dados tem 3 linhas fixas no topo: Linha 1 = cabeçalho "
     "(NUNCA altere), Linha 2 = dica de formato esperado (cinza), Linha 3 = "
     "um EXEMPLO preenchido (amarelo). Comece a digitar seus dados a partir "
     "da linha 4 (na aba Servicos_Equipe, o exemplo ocupa as linhas 3 a 5 — "
     "veja o item 6 abaixo — e os dados começam na linha 6).", False),
    ("3. Apague as linhas de dica e exemplo antes de importar os dados de "
     "verdade, ou apenas avise quem for importar para ignorá-las.", False),
    ("4. A coluna 'processo' é o elo entre as abas Agenda, Clientes, Servicos, "
     "Servicos_Equipe e Avaliacoes_Brutas. Preencha SEMPRE com o mesmo código, "
     "exatamente igual (copie e cole, não redigite), em todas as abas.", False),
    ("5. Ordem recomendada de preenchimento: Agenda primeiro (é ali que o "
     "processo nasce) → Clientes → Servicos → Servicos_Equipe → "
     "Avaliacoes_Brutas.", False),
    ("6. Aba Servicos_Equipe: um serviço pode ter de 3 a 10 pessoas na "
     "equipe. Cada PESSOA é uma LINHA separada (não uma lista numa célula "
     "só). Repita o mesmo 'processo' e 'os' em todas as linhas daquele "
     "serviço, mudando apenas o nome do colaborador. A ordem das pessoas "
     "não importa.", False),
    ("7. Aba Dicionario: guarda as listas de valores padronizados (cidade, "
     "empresa, agente, coordenadora, status, modal, tipo, faturamento, "
     "sim/não). As listas suspensas das outras abas usam essas colunas. "
     "Para adicionar uma opção nova, digite na próxima linha vazia da "
     "coluna correspondente do Dicionario — o dropdown atualiza sozinho.", False),
    ("8. Colunas de data ficam no formato AAAA-MM-DD (ex: 2026-03-15). As "
     "células já vêm formatadas — digite ou use o seletor de data do Excel.", False),
    ("9. Deixe em branco o que você não souber. Não escreva 'N/A', '-', "
     "'não sei' etc. — campo vazio de verdade é mais fácil de tratar depois.", False),
    ("10. Dúvida ou processo fora do padrão: anote na coluna de observações "
     "da aba correspondente (anotacoes_agenda / anotacoes_servico / "
     "comentario) em vez de forçar um valor errado só para preencher a célula.", False),
]
for i, (texto, titulo) in enumerate(instrucoes, start=1):
    celula = ws.cell(row=i, column=1, value=texto)
    if titulo:
        celula.font = Font(name=FONTE, bold=True, size=14, color=AZUL_ESCURO)
    else:
        celula.font = Font(name=FONTE, size=11)
    celula.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 18 if texto else 6


# =====================================================================
# Aba: Dicionario (listas padronizadas que alimentam os dropdowns)
# Título de cada coluna fica na LINHA 2 (não na 1) só para deixar a
# linha 1 livre como um aviso geral da aba.
# =====================================================================
ws = wb.create_sheet("Dicionario")
colunas_dicionario = {
    "A": ("tipo", ["Internacional", "Nacional", "Local"]),
    "B": ("status (agenda e servico)", ["Em andamento", "Concluído", "Cancelado", "Aguardando cliente"]),
    "C": ("modal", ["Marítimo", "Aéreo", "Terrestre", "Multimodal"]),
    "D": ("sim_nao", ["Sim", "Não"]),
    "E": ("faturamento", ["Faturado", "Pendente", "Não faturável"]),
    "F": ("cidade", ["Brasília", "São Paulo", "Rio de Janeiro"]),
    "G": ("empresa", ["Gerson & Grey"]),
    "H": ("agente", []),
    "I": ("coordenadora", []),
}
aviso = ws.cell(row=1, column=1,
                 value="Cada coluna abaixo alimenta uma lista suspensa nas outras abas. "
                       "Adicione valores novos na primeira linha vazia da coluna correspondente.")
aviso.font = Font(name=FONTE, italic=True, size=10, color="595959")
ws.merge_cells("A1:I1")
for letra, (titulo, valores) in colunas_dicionario.items():
    celula_titulo = ws[f"{letra}2"]
    celula_titulo.value = titulo
    celula_titulo.font = Font(name=FONTE, bold=True, color="FFFFFF")
    celula_titulo.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    celula_titulo.alignment = Alignment(horizontal="center")
    ws.column_dimensions[letra].width = 22
    for i, valor in enumerate(valores, start=3):
        ws[f"{letra}{i}"] = valor
ws.freeze_panes = "A3"

# =====================================================================
# Aba: Agenda
# =====================================================================
especificacoes_agenda = [
    ("processo", "Número inteiro (ex.: 15000)", 15007),
    ("recebido_em", "aaaa-mm-dd", "2026-01-10"),
    ("data_a_inicial", "aaaa-mm-dd", "2026-02-01"),
    ("data_b_inicial", "aaaa-mm-dd", "2026-02-15"),
    ("data_a_ofertada", "aaaa-mm-dd", "2026-02-03"),
    ("data_b_ofertada", "aaaa-mm-dd", "2026-02-17"),
    ("volume", "Número decimal, m³ (ex.: 45.5)", 45.5),
    ("engradados", "Número inteiro", 3),
    ("caixas", "Número inteiro", 60),
    ("lifts", "Número inteiro", 1),
    ("tipo", "Selecione da lista (Dicionario)", "Internacional"),
    ("status_agenda", "Selecione da lista (Dicionario)", "Em andamento"),
    ("origem", "Texto: cidade, país", "Brasília, BR"),
    ("destino", "Texto: cidade, país", "Lisboa, PT"),
    ("anotacoes_agenda", "Texto livre", "Cliente pediu prioridade no içamento"),
]
ws, colunas_agenda = montar_aba_dados("Agenda", especificacoes_agenda)
aplicar_formato_data(ws, colunas_agenda, [
    "recebido_em", "data_a_inicial", "data_b_inicial", "data_a_ofertada", "data_b_ofertada"
])
adicionar_dropdown_dicionario(ws, colunas_agenda, "tipo", "A")
adicionar_dropdown_dicionario(ws, colunas_agenda, "status_agenda", "B")

# =====================================================================
# Aba: Clientes
# =====================================================================
especificacoes_clientes = [
    ("processo", "Número inteiro (mesmo processo da aba Agenda)", 15007),
    ("nome_cliente", "Texto: nome completo", "Mariana Costa"),
    ("email", "E-mail", "mariana.costa@email.com"),
    ("email_agendor", "E-mail (sistema Agendor)", "mariana.agendor@email.com"),
    ("agente", "Selecione da lista (Dicionario) ou digite um novo", "Agente Lisboa Mudanças"),
    ("empresa", "Selecione da lista (Dicionario) ou digite um novo", "Gerson & Grey"),
]
ws, colunas_clientes = montar_aba_dados("Clientes", especificacoes_clientes)
adicionar_dropdown_dicionario(ws, colunas_clientes, "agente", "H")
adicionar_dropdown_dicionario(ws, colunas_clientes, "empresa", "G")

# =====================================================================
# Aba: Servicos  (SEM avaliacao e SEM equipe — ver Avaliacoes_Brutas e
# Servicos_Equipe, respectivamente)
# =====================================================================
especificacoes_servicos = [
    ("processo", "Número inteiro (mesmo processo da aba Agenda)", 15007),
    ("servicos", "Texto: descrição do serviço", "Embalagem + Transporte Marítimo"),
    ("modal", "Selecione da lista (Dicionario)", "Marítimo"),
    ("os", "Código da OS (identifica este serviço dentro do processo)", "OS-4451"),
    ("ref_externa", "Texto/código de referência externa", "REF-EXT-991"),
    ("cidade", "Selecione da lista (Dicionario) ou digite uma nova", "Brasília"),
    ("data_inicio", "aaaa-mm-dd", "2026-02-01"),
    ("data_final", "aaaa-mm-dd", "2026-02-14"),
    ("m3_real", "Número decimal, m³ (ex.: 44.8)", 44.8),
    ("quant_itens", "Número inteiro", 120),
    ("peso_kg", "Número decimal, kg", 3200.0),
    ("peso_bruto", "Número decimal, kg", 3400.0),
    ("peso_bruto_real_1", "Número decimal, kg", 3390.0),
    ("peso_liquido", "Número decimal, kg", 3200.0),
    ("peso_liquido_real_1", "Número decimal, kg", 3195.0),
    ("etd", "aaaa-mm-dd", "2026-02-05"),
    ("eta", "aaaa-mm-dd", "2026-03-01"),
    ("liftvan", "Texto/código do liftvan", "LV-002"),
    ("peso_liftvan", "Número decimal, kg", 1800.0),
    ("tara_liftvan", "Número decimal, kg", 350.0),
    ("icamento", "Texto: tipo de içamento", "Guindaste"),
    ("container_20", "Selecione da lista (Dicionario): Sim/Não", "Não"),
    ("container_40", "Selecione da lista (Dicionario): Sim/Não", "Sim"),
    ("quant_container_20", "Número inteiro", 0),
    ("quant_cont_40", "Número inteiro", 1),
    ("contents", "Texto: conteúdo transportado", "Móveis e utensílios domésticos"),
    ("empresa", "Selecione da lista (Dicionario) ou digite uma nova", "Gerson & Grey"),
    ("tipo_cliente", "Texto: ex. Residencial, Corporativo", "Residencial"),
    ("status_servico", "Selecione da lista (Dicionario)", "Concluído"),
    ("faturamento", "Selecione da lista (Dicionario)", "Faturado"),
    ("fatura", "Texto/número da nota fiscal", "NF-8821"),
    ("coordenadora", "Selecione da lista (Dicionario) ou digite uma nova", "Ana Paula"),
    ("anotacoes_servico", "Texto livre", "Cliente elogiou pontualidade da equipe"),
]
ws, colunas_servicos = montar_aba_dados("Servicos", especificacoes_servicos)
aplicar_formato_data(ws, colunas_servicos, ["data_inicio", "data_final", "etd", "eta"])
adicionar_dropdown_dicionario(ws, colunas_servicos, "modal", "C")
adicionar_dropdown_dicionario(ws, colunas_servicos, "container_20", "D")
adicionar_dropdown_dicionario(ws, colunas_servicos, "container_40", "D")
adicionar_dropdown_dicionario(ws, colunas_servicos, "status_servico", "B")
adicionar_dropdown_dicionario(ws, colunas_servicos, "faturamento", "E")
adicionar_dropdown_dicionario(ws, colunas_servicos, "cidade", "F")
adicionar_dropdown_dicionario(ws, colunas_servicos, "empresa", "G")
adicionar_dropdown_dicionario(ws, colunas_servicos, "coordenadora", "I")

# =====================================================================
# Aba: Servicos_Equipe (nova — normaliza o antigo campo de texto "equipe")
# Uma LINHA por PESSOA por serviço. 3 linhas de exemplo (não 1), porque
# uma equipe de verdade tem várias pessoas — isso demonstra o padrão de
# "uma linha por pessoa" mais claramente que um único exemplo isolado.
# =====================================================================
colunas_equipe = ["processo", "os", "nome_colaborador"]
dicas_equipe = [
    "Número inteiro (mesmo processo da aba Agenda)",
    "Código da OS (mesmo da aba Servicos)",
    "Nome completo — uma pessoa por linha",
]
exemplos_equipe = [
    [15007, "OS-4451", "Ana Paula"],
    [15007, "OS-4451", "Bruno Nogueira"],
    [15007, "OS-4451", "Carlos Eduardo"],
]
ws = wb.create_sheet("Servicos_Equipe")
ws.append(colunas_equipe)
ws.append(dicas_equipe)
for linha_exemplo in exemplos_equipe:
    ws.append(linha_exemplo)
estilizar_cabecalho(ws, len(colunas_equipe))
estilizar_linha_dica(ws, len(colunas_equipe))
for linha in (3, 4, 5):
    estilizar_linha_exemplo(ws, len(colunas_equipe), linha=linha)
for i, _ in enumerate(colunas_equipe, start=1):
    ws.column_dimensions[get_column_letter(i)].width = 22
ws.freeze_panes = "A6"  # aqui o exemplo ocupa as linhas 3 a 5; dados começam na 6
# Nenhuma coluna desta aba é categórica (processo, os e nome_colaborador
# são todos texto/número livre), então não há dropdown do Dicionario aqui.

# =====================================================================
# Aba: Avaliacoes_Brutas
# =====================================================================
especificacoes_avaliacoes = [
    ("processo", "Número inteiro (mesmo processo da aba Agenda)", 15007),
    ("data", "aaaa-mm-dd", "2026-02-16"),
    ("ano", "aaaa", 2026),
    ("mes", "Número de 1 a 12", 2),
    ("nota_pontualidade_coord", "Número inteiro de 0 a 10", 10),
    ("nota_limpeza_embalagem", "Número inteiro de 0 a 10", 9),
    ("nota_cortesia_carregamento", "Número inteiro de 0 a 10", 10),
    ("nota_tecnica_cortesia", "Número inteiro de 0 a 10", 9),
    ("comentario", "Texto livre", "Excelente atendimento da coordenadora"),
]
ws, colunas_avaliacoes = montar_aba_dados("Avaliacoes_Brutas", especificacoes_avaliacoes)
aplicar_formato_data(ws, colunas_avaliacoes, ["data"])

# =====================================================================
# Ordena as abas na ordem lógica de preenchimento
# =====================================================================
ordem_final = [
    "Instrucoes", "Dicionario", "Agenda", "Clientes",
    "Servicos", "Servicos_Equipe", "Avaliacoes_Brutas",
]
wb._sheets = [wb[nome] for nome in ordem_final]

wb.save("template_migracao_dados.xlsx")
print("Planilha gerada: template_migracao_dados.xlsx")
